
# ===============================================
# PremierNiK Fantasy - Colab V2.1 (Incremental) + Validation Alerts
# ===============================================
# - Reads Teams.xlsx (sheet "Teams": Team, Player 1 ID..Player 4 ID, Captain ID)
# - Reads Results_V2.xlsx (sheets: Calendar, Fixtures, Standings)
# - Input target GW and optional force rebuild
# - Incremental update computes only missing GWs; force rebuild recomputes 1..GW
# - Auto-detect BEST (✅) and WORST (❌) per GW
# - Auto-update Standings (Wins/Draws/Losses, Bests, Worsts, point, Total Point, Diff+/-)
# - English validation alerts for missing/mismatched team names
# - Creates a backup of Results_V2.xlsx before saving updated file
# ===============================================

import os
import shutil
import pandas as pd
import requests
from datetime import datetime
from openpyxl import load_workbook

# Colab file upload support (optional)
try:
    from google.colab import files
    IN_COLAB = True
except Exception:
    IN_COLAB = False

# ---------------- Helpers ----------------
def backup_file(path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}_backup_{ts}{ext}"
    shutil.copy(path, backup_path)
    return backup_path

def normalize_name(name):
    return str(name).strip()

def read_excel(path_or_bytes, sheet=None):
    return pd.read_excel(path_or_bytes, sheet_name=sheet)

def get_player_entry_points(team_id, gw):
    """Return entry_history.points for the entry (already net of transfer hits)."""
    url = f"https://fantasy.premierleague.com/api/entry/{int(team_id)}/event/{int(gw)}/picks/"
    r = requests.get(url, timeout=20)
    if r.status_code != 200:
        raise RuntimeError(f"FPL API error for team {team_id}, GW{gw}: HTTP {r.status_code}")
    data = r.json()
    points = data.get("entry_history", {}).get("points", 0)
    return int(points)

def compute_group_points_for_gw(team_row, gw):
    """Sum points for 4 players + add captain's points again to double him at league level."""
    pids = []
    for col in ["Player 1 ID", "Player 2 ID", "Player 3 ID", "Player 4 ID"]:
        pid = team_row.get(col)
        if pd.isna(pid) or str(pid).strip() == "":
            continue
        pids.append(int(float(pid)))
    if len(pids) == 0:
        return 0
    cap = team_row.get("Captain ID")
    cap = None if (pd.isna(cap) or str(cap).strip() == "") else int(float(cap))

    total = 0
    cap_pts = 0
    for pid in pids:
        pts = get_player_entry_points(pid, gw)
        total += pts
        if cap is not None and pid == cap:
            cap_pts = pts
    return total + cap_pts  # double captain at league level

def ensure_sheet_headers(ws, headers):
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1,1).value is None:
        ws.append(headers)

def recompute_standings(wb):
    ws = wb["Fixtures"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    rows = []
    for i in range(2, ws.max_row+1):
        rows.append({headers[c-1]: ws.cell(i, c).value for c in range(1, ws.max_column+1)})
    if not rows:
        # clear standings
        ws_s = wb["Standings"]
        ws_s.delete_rows(2, ws_s.max_row)
        return

    fx = pd.DataFrame(rows)
    fx["Team"] = fx["Team"].apply(normalize_name)
    fx["Opponent"] = fx["Opponent"].apply(normalize_name)
    fx["Points For"] = pd.to_numeric(fx["Points For"], errors="coerce").fillna(0).astype(int)
    fx["Points Against"] = pd.to_numeric(fx["Points Against"], errors="coerce").fillna(0).astype(int)
    fx["Result"] = fx["Result"].fillna("")
    fx["BEST"] = fx["BEST"].fillna("")
    fx["WORST"] = fx["WORST"].fillna("")

    agg = fx.groupby("Team").agg(
        Match_Played=("GW", "count"),
        Wins=("Result", lambda s: (s=="Win").sum()),
        Draws=("Result", lambda s: (s=="Draw").sum()),
        Losses=("Result", lambda s: (s=="Loss").sum()),
        Bests=("BEST", lambda s: (s=="✅").sum()),
        Worsts=("WORST", lambda s: (s=="❌").sum()),
        Total_For=("Points For", "sum"),
        Total_Against=("Points Against", "sum")
    ).reset_index()

    agg["Diff+/-"] = agg["Total_For"] - agg["Total_Against"]
    agg["point"] = (agg["Wins"]*3 + agg["Draws"]*1 + agg["Bests"]*1 - agg["Worsts"]*1)
    agg["Total Point"] = agg["Total_For"]

    # Order & sort
    agg = agg.rename(columns={"Match_Played":"Match Played", "Total Point":"Total Point"})
    agg = agg[["Team","Match Played","Wins","Draws","Losses","Bests","Worsts","point","Total Point","Diff+/-"]]
    agg = agg.sort_values(["point","Diff+/-","Total Point"], ascending=[False, False, False])

    ws_s = wb["Standings"]
    ws_s.delete_rows(2, ws_s.max_row)
    for _, r in agg.iterrows():
        ws_s.append(list(r.values))

def infer_best_worst(rows_for_gw):
    if not rows_for_gw:
        return set(), set()
    max_pts = max(r["Points For"] for r in rows_for_gw)
    min_pts = min(r["Points For"] for r in rows_for_gw)
    best = {r["Team"] for r in rows_for_gw if r["Points For"] == max_pts}
    worst = {r["Team"] for r in rows_for_gw if r["Points For"] == min_pts}
    return best, worst

# --------------- Main ---------------
def main():
    print("=== PremierNiK Fantasy Colab V2.1 (Incremental) + Validation ===")
    # Upload or use local paths
    if IN_COLAB:
        print("📂 Upload Teams.xlsx")
        up_t = files.upload()
        teams_path = list(up_t.keys())[0]
        print("📂 Upload Results_V2.xlsx")
        up_r = files.upload()
        results_path = list(up_r.keys())[0]
    else:
        teams_path = "Teams.xlsx"
        results_path = "Results_V2.xlsx"
        if not (os.path.exists(teams_path) and os.path.exists(results_path)):
            raise SystemExit("Place Teams.xlsx and Results_V2.xlsx next to this script and rerun.")

    # Inputs
    try:
        target_gw = int(input("Enter target GW (e.g., 10): ").strip())
    except:
        raise SystemExit("Invalid GW number.")
    force = input("Force rebuild from GW1? (y/n): ").strip().lower().startswith("y")

    # Read Results workbook
    wb = load_workbook(results_path)
    for sheet in ["Calendar","Fixtures","Standings"]:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"Missing sheet '{sheet}' in Results_V2.xlsx")

    # Read Calendar
    cal_df = read_excel(results_path, sheet="Calendar")
    cal_df = cal_df.dropna(subset=["GW","Team","Opponent"])
    cal_df["GW"] = pd.to_numeric(cal_df["GW"], errors="coerce").fillna(0).astype(int)
    cal_df["Team"] = cal_df["Team"].apply(normalize_name)
    cal_df["Opponent"] = cal_df["Opponent"].apply(normalize_name)
    cal_df = cal_df[cal_df["GW"].between(1, target_gw)]

    # Read Teams sheet explicitly
    teams_df = pd.read_excel(teams_path, sheet_name="Teams")
    # Clean header names from hidden chars/spaces
    teams_df.columns = [str(c).strip().replace('\u200f','').replace('\u200e','') for c in teams_df.columns]
    if "Team" not in teams_df.columns:
        raise SystemExit(
            "The 'Team' column is missing in the 'Teams' sheet.\n"
            f"Found columns: {list(teams_df.columns)}\n"
            "Make sure cell A1 is exactly 'Team' (no extra spaces/merged rows)."
        )
    teams_df["Team"] = teams_df["Team"].apply(normalize_name)

    # ---------- Validation of team names ----------
    calendar_teams = set(pd.concat([cal_df["Team"], cal_df["Opponent"]]).unique())
    teams_master = set(teams_df["Team"].unique())
    missing = sorted(list(calendar_teams - teams_master))
    if missing:
        msg_lines = [
            "Validation error: Some team names in 'Calendar' were NOT found in 'Teams.xlsx'.",
            "Please fix the spelling to match exactly (case/spacing) and try again.",
            "Missing team names:"
        ] + [f"- {t}" for t in missing]
        raise SystemExit("\n".join(msg_lines))
    # ---------------------------------------------

    ws_fx = wb["Fixtures"]
    fx_headers = [ws_fx.cell(1, c).value for c in range(1, ws_fx.max_column+1)]
    fx_rows = []
    for i in range(2, ws_fx.max_row+1):
        fx_rows.append({fx_headers[c-1]: ws_fx.cell(i, c).value for c in range(1, ws_fx.max_column+1)})
    fixtures_df = pd.DataFrame(fx_rows) if fx_rows else pd.DataFrame(columns=fx_headers)

    # Decide range
    if force:
        compute_from = 1
        ws_fx.delete_rows(2, ws_fx.max_row)
        ws_s = wb["Standings"]
        ws_s.delete_rows(2, ws_s.max_row)
    else:
        last_gw = 0 if fixtures_df.empty else int(fixtures_df["GW"].max())
        compute_from = last_gw + 1

    if compute_from > target_gw:
        print(f"Nothing to compute. Last saved GW >= target GW ({compute_from-1} >= {target_gw}).")
        recompute_standings(wb)
        out_name = "Results_V2_updated.xlsx"
        backup_file(results_path)
        wb.save(out_name)
        if IN_COLAB:
            files.download(out_name)
        return

    # Lookup teams
    teams_lookup = {normalize_name(r["Team"]): r for _, r in teams_df.iterrows()}

    rows_to_append = []
    for gw in range(compute_from, target_gw+1):
        cal_gw = cal_df[cal_df["GW"] == gw].copy()
        if cal_gw.empty:
            print(f"⚠ Calendar has no rows for GW{gw}. Skipping.")
            continue

        # Compute "Points For" per team
        per_team_points = {}
        for _, row in cal_gw.iterrows():
            team_name = row["Team"]
            team_row = teams_lookup.get(team_name)
            if team_row is None:
                raise SystemExit(f"Validation error: Team '{team_name}' (GW {gw}) not present in Teams.xlsx.")
            pts = compute_group_points_for_gw(team_row, gw)
            per_team_points[team_name] = int(pts)

        # Build fixtures rows (one line per team)
        rows_gw = []
        for _, row in cal_gw.iterrows():
            team = row["Team"]
            opp = row["Opponent"]
            pf = per_team_points.get(team, 0)
            pa = per_team_points.get(opp, 0)
            result = "Win" if pf > pa else ("Draw" if pf == pa else "Loss")
            rows_gw.append({
                "GW": gw, "Team": team, "Opponent": opp,
                "Points For": pf, "Points Against": pa,
                "Result": result, "BEST": "", "WORST": ""
            })

        # BEST/WORST auto marks
        best_set, worst_set = infer_best_worst(rows_gw)
        for r in rows_gw:
            if r["Team"] in best_set:
                r["BEST"] = "✅"
            if r["Team"] in worst_set:
                r["WORST"] = "❌"

        # Stage rows for append
        rows_to_append.extend([[
            r["GW"], r["Team"], r["Opponent"], r["Points For"], r["Points Against"],
            r["Result"], r["BEST"], r["WORST"]
        ] for r in rows_gw])

    # Append to Fixtures
    for row in rows_to_append:
        ws_fx.append(row)

    # Recompute standings and save
    recompute_standings(wb)
    out_name = "Results_V2_updated.xlsx"
    backup = backup_file(results_path)
    wb.save(out_name)
    print(f"✅ Done. Backup saved as: {backup}")
    print(f"➡ Updated file saved as: {out_name}")
    if IN_COLAB:
        files.download(out_name)

if __name__ == "__main__":
    main()
